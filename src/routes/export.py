from flask import Blueprint, request, jsonify, send_file, render_template_string, current_app
from src.models.database import db
from src.models.models import Process, Bid, Supplier, Document, Alert, EvaluationCriteria, BidEvaluation, BidRanking
from datetime import datetime
import tempfile
import os
import logging
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
import io
import base64

logger = logging.getLogger(__name__)
export_bp = Blueprint('export', __name__)

# Template HTML para PDF
PDF_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Informe de Proceso - {{process.process_number}}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .header { text-align: center; border-bottom: 2px solid #333; padding-bottom: 20px; margin-bottom: 30px; }
        .section { margin-bottom: 30px; }
        .section h2 { color: #333; border-bottom: 1px solid #ccc; padding-bottom: 5px; }
        .info-table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
        .info-table th, .info-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        .info-table th { background-color: #f2f2f2; font-weight: bold; }
        .status-active { color: #28a745; font-weight: bold; }
        .status-completed { color: #007bff; font-weight: bold; }
        .status-cancelled { color: #dc3545; font-weight: bold; }
        .ranking-winner { background-color: #d4edda; }
        .footer { margin-top: 50px; text-align: center; font-size: 12px; color: #666; }
    </style>
</head>
<body>
    <div class="header">
        <h1>Informe de Proceso de Compra</h1>
        <h2>{{process.process_number}} - {{process.title}}</h2>
        <p>Generado el {{current_date}}</p>
    </div>

    <div class="section">
        <h2>Información General</h2>
        <table class="info-table">
            <tr><th>Número de Proceso</th><td>{{process.process_number}}</td></tr>
            <tr><th>Título</th><td>{{process.title}}</td></tr>
            <tr><th>Descripción</th><td>{{process.description or 'N/A'}}</td></tr>
            <tr><th>Tipo</th><td>{{process_type_label}}</td></tr>
            <tr><th>Estado</th><td class="status-{{process.status}}">{{process.status}}</td></tr>
            <tr><th>Presupuesto</th><td>{{budget_formatted}}</td></tr>
            <tr><th>Fecha de Inicio</th><td>{{start_date_formatted}}</td></tr>
            <tr><th>Fecha de Fin</th><td>{{end_date_formatted}}</td></tr>
            <tr><th>Fecha de Creación</th><td>{{created_date_formatted}}</td></tr>
        </table>
    </div>

    {% if bids %}
    <div class="section">
        <h2>Ofertas Recibidas ({{bids|length}})</h2>
        <table class="info-table">
            <thead>
                <tr>
                    <th>Proveedor</th>
                    <th>Monto</th>
                    <th>Puntaje Total</th>
                    <th>Estado</th>
                    <th>Fecha de Envío</th>
                </tr>
            </thead>
            <tbody>
                {% for bid in bids %}
                <tr>
                    <td>{{bid.supplier_name}}</td>
                    <td>{{bid.bid_amount_formatted}}</td>
                    <td>{{bid.total_score or 'N/A'}}</td>
                    <td>{{bid.status}}</td>
                    <td>{{bid.submission_date_formatted}}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% endif %}

    {% if ranking %}
    <div class="section">
        <h2>Ranking Final</h2>
        <table class="info-table">
            <thead>
                <tr>
                    <th>Posición</th>
                    <th>Proveedor</th>
                    <th>Monto</th>
                    <th>Puntaje Total</th>
                    <th>Recomendación</th>
                </tr>
            </thead>
            <tbody>
                {% for rank in ranking %}
                <tr {% if rank.ranking_position == 1 %}class="ranking-winner"{% endif %}>
                    <td>{{rank.ranking_position}}</td>
                    <td>{{rank.supplier_name}}</td>
                    <td>{{rank.bid_amount_formatted}}</td>
                    <td>{{rank.weighted_total_score}}</td>
                    <td>{{rank.recommendation}}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% endif %}

    {% if criteria %}
    <div class="section">
        <h2>Criterios de Evaluación</h2>
        <table class="info-table">
            <thead>
                <tr>
                    <th>Criterio</th>
                    <th>Tipo</th>
                    <th>Peso (%)</th>
                    <th>Puntaje Máximo</th>
                </tr>
            </thead>
            <tbody>
                {% for criterion in criteria %}
                <tr>
                    <td>{{criterion.name}}</td>
                    <td>{{criterion.criteria_type}}</td>
                    <td>{{criterion.weight}}%</td>
                    <td>{{criterion.max_score}}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% endif %}

    {% if documents %}
    <div class="section">
        <h2>Documentos Asociados ({{documents|length}})</h2>
        <table class="info-table">
            <thead>
                <tr>
                    <th>Nombre</th>
                    <th>Tipo</th>
                    <th>Tamaño</th>
                    <th>Fecha de Subida</th>
                </tr>
            </thead>
            <tbody>
                {% for doc in documents %}
                <tr>
                    <td>{{doc.original_filename}}</td>
                    <td>{{doc.document_type}}</td>
                    <td>{{doc.file_size_formatted}}</td>
                    <td>{{doc.upload_date_formatted}}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% endif %}

    <div class="footer">
        <p>Sistema de Gestión de Compras y Licitaciones</p>
        <p>Informe generado automáticamente</p>
    </div>
</body>
</html>
"""

@export_bp.route('/process/<int:process_id>/pdf', methods=['GET'])
def export_process_pdf(process_id):
    """Exportar informe completo de proceso a PDF"""
    try:
        # Obtener datos del proceso
        process = Process.query.get_or_404(process_id)
        bids = Bid.query.filter_by(process_id=process_id).all()
        documents = Document.query.filter_by(process_id=process_id).all()
        criteria = EvaluationCriteria.query.filter_by(process_id=process_id).all()
        ranking = BidRanking.query.filter_by(process_id=process_id)\
                                 .order_by(BidRanking.ranking_position).all()
        
        # Preparar datos para el template
        template_data = {
            'process': process,
            'current_date': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'process_type_label': 'Compra Simple' if process.process_type == 'simple_purchase' else 'Licitación Grande',
            'budget_formatted': format_currency(process.budget) if process.budget else 'N/A',
            'start_date_formatted': process.start_date.strftime('%d/%m/%Y') if process.start_date else 'N/A',
            'end_date_formatted': process.end_date.strftime('%d/%m/%Y') if process.end_date else 'N/A',
            'created_date_formatted': process.created_date.strftime('%d/%m/%Y') if process.created_date else 'N/A',
            'bids': [],
            'documents': [],
            'criteria': [c.to_dict() for c in criteria],
            'ranking': []
        }
        
        # Formatear ofertas
        for bid in bids:
            bid_data = bid.to_dict()
            bid_data['bid_amount_formatted'] = format_currency(bid.bid_amount) if bid.bid_amount else 'N/A'
            bid_data['submission_date_formatted'] = bid.submission_date.strftime('%d/%m/%Y') if bid.submission_date else 'N/A'
            template_data['bids'].append(bid_data)
        
        # Formatear documentos
        for doc in documents:
            doc_data = doc.to_dict()
            doc_data['file_size_formatted'] = format_file_size(doc.file_size) if doc.file_size else 'N/A'
            doc_data['upload_date_formatted'] = doc.upload_date.strftime('%d/%m/%Y') if doc.upload_date else 'N/A'
            template_data['documents'].append(doc_data)
        
        # Formatear ranking
        for rank in ranking:
            rank_data = rank.to_dict()
            rank_data['bid_amount_formatted'] = format_currency(rank.bid_amount) if rank.bid_amount else 'N/A'
            template_data['ranking'].append(rank_data)
        
        # Renderizar HTML
        from jinja2 import Template
        template = Template(PDF_TEMPLATE)
        html_content = template.render(**template_data)
        
        # Generar PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            pisa_status = pisa.CreatePDF(html_content, dest=tmp_file)
            
            if pisa_status.err:
                return jsonify({'error': 'Error generando PDF'}), 500
            
            tmp_file_path = tmp_file.name
        
        filename = f"proceso_{process.process_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        logger.error(f"Error exporting process PDF: {str(e)}")
        return jsonify({'error': str(e)}), 500

@export_bp.route('/process/<int:process_id>/excel', methods=['GET'])
def export_process_excel(process_id):
    """Exportar informe completo de proceso a Excel"""
    try:
        # Obtener datos del proceso
        process = Process.query.get_or_404(process_id)
        bids = Bid.query.filter_by(process_id=process_id).all()
        documents = Document.query.filter_by(process_id=process_id).all()
        criteria = EvaluationCriteria.query.filter_by(process_id=process_id).all()
        ranking = BidRanking.query.filter_by(process_id=process_id)\
                                 .order_by(BidRanking.ranking_position).all()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Información General
        ws1 = wb.active
        ws1.title = "Información General"
        
        # Título
        ws1['A1'] = f"Informe de Proceso: {process.process_number}"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:B1')
        
        # Información del proceso
        info_data = [
            ['Número de Proceso', process.process_number],
            ['Título', process.title],
            ['Descripción', process.description or 'N/A'],
            ['Tipo', 'Compra Simple' if process.process_type == 'simple_purchase' else 'Licitación Grande'],
            ['Estado', process.status],
            ['Presupuesto', process.budget or 'N/A'],
            ['Fecha de Inicio', process.start_date.strftime('%d/%m/%Y') if process.start_date else 'N/A'],
            ['Fecha de Fin', process.end_date.strftime('%d/%m/%Y') if process.end_date else 'N/A'],
            ['Fecha de Creación', process.created_date.strftime('%d/%m/%Y') if process.created_date else 'N/A'],
            ['Generado el', datetime.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            ws1[f'A{i}'] = label
            ws1[f'B{i}'] = value
            ws1[f'A{i}'].font = Font(bold=True)
            ws1[f'A{i}'].border = border
            ws1[f'B{i}'].border = border
        
        # Ajustar ancho de columnas
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 40
        
        # Hoja 2: Ofertas
        if bids:
            ws2 = wb.create_sheet(title="Ofertas")
            
            # Headers
            headers = ['Proveedor', 'Monto', 'Puntaje Total', 'Estado', 'Fecha de Envío']
            for col, header in enumerate(headers, start=1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Datos
            for row, bid in enumerate(bids, start=2):
                ws2.cell(row=row, column=1, value=bid.supplier.name if bid.supplier else 'N/A').border = border
                ws2.cell(row=row, column=2, value=bid.bid_amount or 0).border = border
                ws2.cell(row=row, column=3, value=bid.total_score or 'N/A').border = border
                ws2.cell(row=row, column=4, value=bid.status).border = border
                ws2.cell(row=row, column=5, value=bid.submission_date.strftime('%d/%m/%Y') if bid.submission_date else 'N/A').border = border
            
            # Ajustar ancho de columnas
            for col in range(1, 6):
                ws2.column_dimensions[chr(64 + col)].width = 15
        
        # Hoja 3: Ranking
        if ranking:
            ws3 = wb.create_sheet(title="Ranking")
            
            # Headers
            headers = ['Posición', 'Proveedor', 'Monto', 'Puntaje Técnico', 'Puntaje Comercial', 'Puntaje Financiero', 'Puntaje Total', 'Recomendación']
            for col, header in enumerate(headers, start=1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Datos
            for row, rank in enumerate(ranking, start=2):
                ws3.cell(row=row, column=1, value=rank.ranking_position).border = border
                ws3.cell(row=row, column=2, value=rank.supplier_name).border = border
                ws3.cell(row=row, column=3, value=rank.bid_amount or 0).border = border
                ws3.cell(row=row, column=4, value=rank.technical_score or 0).border = border
                ws3.cell(row=row, column=5, value=rank.commercial_score or 0).border = border
                ws3.cell(row=row, column=6, value=rank.financial_score or 0).border = border
                ws3.cell(row=row, column=7, value=rank.weighted_total_score).border = border
                ws3.cell(row=row, column=8, value=rank.recommendation).border = border
                
                # Resaltar ganador
                if rank.ranking_position == 1:
                    for col in range(1, 9):
                        ws3.cell(row=row, column=col).fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            
            # Ajustar ancho de columnas
            for col in range(1, 9):
                ws3.column_dimensions[chr(64 + col)].width = 12
        
        # Hoja 4: Criterios de Evaluación
        if criteria:
            ws4 = wb.create_sheet(title="Criterios")
            
            # Headers
            headers = ['Criterio', 'Tipo', 'Peso (%)', 'Puntaje Máximo', 'Descripción']
            for col, header in enumerate(headers, start=1):
                cell = ws4.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Datos
            for row, criterion in enumerate(criteria, start=2):
                ws4.cell(row=row, column=1, value=criterion.name).border = border
                ws4.cell(row=row, column=2, value=criterion.criteria_type).border = border
                ws4.cell(row=row, column=3, value=criterion.weight).border = border
                ws4.cell(row=row, column=4, value=criterion.max_score).border = border
                ws4.cell(row=row, column=5, value=criterion.description or 'N/A').border = border
            
            # Ajustar ancho de columnas
            ws4.column_dimensions['A'].width = 20
            ws4.column_dimensions['B'].width = 12
            ws4.column_dimensions['C'].width = 10
            ws4.column_dimensions['D'].width = 12
            ws4.column_dimensions['E'].width = 30
        
        # Guardar archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            wb.save(tmp_file.name)
            tmp_file_path = tmp_file.name
        
        filename = f"proceso_{process.process_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            tmp_file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error exporting process Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

def format_currency(amount):
    """Formatear moneda"""
    if amount is None:
        return 'N/A'
    return f"${amount:,.0f}"

def format_file_size(size_bytes):
    """Formatear tamaño de archivo"""
    if size_bytes is None:
        return 'N/A'
    
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

